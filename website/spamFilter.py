from builtins import str
from builtins import range
import openpyxl
import numpy as np
from django.conf import settings
from .cleanText import clean_string
from sklearn.svm import LinearSVC
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from website.models import Question, Answer

# Get the original dataset
def store():

    # Add data from Excel file
    file_location = settings.PROJECT_DIR + '/DataSet.xlsx'
    workBookOld = openpyxl.load_workbook(file_location)
    dataSheetOld = workBookOld['Data set']

    xData = []
    yData = []

    rows = dataSheetOld.max_row

    for i in range(2, rows+1):

        if (str(dataSheetOld.cell(row = i, column = 2).value) != 'None'):
            xData.append(str(clean_string(dataSheetOld.cell(row = i, column = 1).value)))
            if (str(dataSheetOld.cell(row = i, column = 2).value) == "1"):
                yData.append(1)
            else:
                yData.append(0)

    # Add data from forum questions
    for question in Question.objects.all():
        xData.append(str(clean_string(question.body)))
        if (question.is_spam):
            yData.append(1)
        else:
            yData.append(0)

    # Add data from forum answers
    for answer in Answer.objects.all():
        xData.append(str(clean_string(answer.body)))
        if (answer.is_spam):
            yData.append(1)
        else:
            yData.append(0)

    return xData, yData

    # # NOTE: to train data on the entire dataset, simply return xData and yData
    # # Splitting the data like this is to obtain test cases and calculate the F-score of the learning algorithm
    # xTrain, xTest, yTrain, yTest = train_test_split(xData, yData, test_size = 0.2, random_state = 42)
    # return xTrain, xTest, yTrain, yTest

# Train the data when server checks happening
def train():

    print("Training spam filter...")

    # Create training data
    xTrain, yTrain = store()
    global vectorizer
    xTrainMatrix = vectorizer.fit_transform(xTrain)
    yTrainMatrix = np.asarray(yTrain)

    global model
    model.fit(xTrainMatrix, yTrainMatrix)


# Calculating the F-score
def calc_f_score(xTest, yTest, model, vectorizer):

    xTestMatrix = vectorizer.transform(xTest)
    yTestMatrix = np.asarray(yTest)

    result = model.predict(xTestMatrix)
    matrix = confusion_matrix(yTestMatrix, result)

    fScore = f1_score(yTestMatrix, result, pos_label = 0)
    precision = precision_score(yTestMatrix, result, pos_label = 0)
    recall = recall_score(yTestMatrix, result, pos_label = 0)
    return fScore, precision, recall, matrix

# Test new data for Spam
def predict(emailBody):

    string = clean_string(emailBody)
    if ('httpaddr' in string or 'linktag' in string):
        return "Spam"

    global vectorizer
    featureMatrix = vectorizer.transform([string])
    global model
    result = model.predict(featureMatrix)

    if (1 in result):
        return "Spam"
    else:
        return "Not Spam"

model = LinearSVC(class_weight = 'balanced')
vectorizer = TfidfVectorizer(stop_words = 'english', max_df = 75)
train()
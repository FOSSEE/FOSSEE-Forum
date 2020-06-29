# standard library
from builtins import str, zip
from datetime import date, datetime

# third-party
import openpyxl

# Django
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, User
from django.core import mail
from django.core.mail import EmailMultiAlternatives
from django.db.models import Q
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.template.context_processors import csrf
from django.template.loader import render_to_string
from django.urls import Resolver404, resolve
from django.utils.html import strip_tags
from django.views.decorators.csrf import csrf_exempt

# local Django
from .auto_mail_send import Cron
from .decorators import check_recaptcha
from .forms import AnswerCommentForm, AnswerQuestionForm, NewQuestionForm
from .models import (
    Answer, AnswerComment, FossCategory, ModeratorGroup,
    Notification, Question, Scheduled_Auto_Mail, SubFossCategory,
)
from .spamFilter import predict, train
from .templatetags.helpers import prettify

User = get_user_model()
admins = User.objects.filter(is_superuser=True).values_list('id')


# NON-VIEWS FUNCTIONS

def account_credentials_defined(user):
    """
    Return True if the user has completed his Profile OR
    if the user is a Moderator (Moderators don't need to complete their Profile)
    """
    return ((user.first_name and user.last_name) or is_moderator(user))


def is_moderator(user, question=None):
    """
    Return True if the user is a moderator of the category
    to which the question belongs, if question is provided.
    Return True if the user is a moderator of any category,
    if question is not provided.
    """
    if question:
        return (user.groups.filter(name="forum_moderator").exists() or
                user.groups.filter(moderatorgroup=get_object_or_404(
                    ModeratorGroup, category=question.category)).exists())
    return user.groups.count() > 0


def to_uids(question):
    """
    Return a set of user ids of all the people linked to the Question,
    i.e., User IDs of Question, Answers and Comments' Authors.
    """
    mail_uids = [question.user.id]
    answers = Answer.objects.filter(question_id=question.id,
                                    is_active=True).distinct()
    for answer in answers:
        for comment in AnswerComment.objects.values('uid').filter(
                answer=answer, is_active=True).distinct():
            mail_uids.append(comment['uid'])
        mail_uids.append(answer.uid)
    mail_uids = set(mail_uids)
    return mail_uids


def mod_uids(question=None):
    """
    Return a list of all moderators of the Question, is question is provided.
    Return a list of all moderators on the forum, otherwise.
    """
    if question:
        mods = User.objects.filter(
            Q(groups__moderatorgroup__category=question.category) |
            Q(groups__name="forum_moderator")
        )
    else:
        mods = User.objects.filter(groups__isnull=False)
    uids = []
    for mod in mods:
        uids.append(mod.id)
    return uids


def get_user_email(uid):
    user = User.objects.get(id=uid)
    return user.email


def send_email(subject, plain_message, html_message, from_email, to,
               bcc=None, cc=None, reply_to=None):
    """
    Send Emails to Everyone in the 'to' list, 'bcc' list, and 'cc' list.
    The Email IDs of everyone in the 'to' list and 'cc' list will be
    visible to the recipents of the email.
    """
    if bcc is None:
        bcc = [settings.BCC_EMAIL_ID]
    else:
        bcc.append(settings.BCC_EMAIL_ID)

    email = EmailMultiAlternatives(
        subject,
        plain_message,
        from_email,
        to,
        bcc=bcc,
        cc=cc,
        reply_to=reply_to,
        headers={"Content-type": "text/html;charset=iso-8859-1"},
    )
    email.attach_alternative(html_message, "text/html")
    # if you are sending attachment with content id,
    # subtype must be 'related'.
    # email.mixed_subtype = 'related'
    email.send(fail_silently=True)


def send_email_as_to(subject, plain_message, html_message,
                     from_email, to, reply_to=None):
    """Send Emails to everyone in the 'to' list individually."""
    # The connection to be used for sending the emails so that
    # all the emails can be sent by opening a single connection.
    connection = mail.get_connection(fail_silently=True)
    messages = []
    to.append(settings.BCC_EMAIL_ID)
    for to_email in to:
        email = EmailMultiAlternatives(
            subject,
            plain_message,
            from_email,
            [to_email],
            reply_to=reply_to,
            headers={"Content-type": "text/html;charset=iso-8859-1"},
        )
        email.attach_alternative(html_message, "text/html")
        messages.append(email)
    connection.send_messages(messages)


def can_delete_comment(answer, comment_id):
    """Return True if there are no active comments after the comment to be
    deleted.
    """
    comments = answer.answercomment_set.filter(is_active=True).all()
    for c in comments:
        if c.id > int(comment_id):
            return False
    return True


def process_Spam(content_body, is_spam):
    """
    Update the value of is_spam in DataSet if the question_body already exists
    in DataSet. Add the question body and the corresponding value of is_spam in
    DataSet, otherwise.
    """
    file_location = settings.BASE_DIR + '/Spam_Filter_Data/DataSet.xlsx'
    xfile = openpyxl.load_workbook(file_location)
    sheet = xfile['Data set']
    n = len(sheet['A']) + 1
    for i in range(2, n):
        if(content_body == str(sheet.cell(row=i, column=1).value)):
            sheet.cell(row=i, column=2).value = is_spam
            xfile.save(file_location)
            return
    sheet['A%s' % n] = content_body
    sheet['B%s' % n] = is_spam
    xfile.save(file_location)


def send_remider_mail():
    if date.today().weekday() == 1 or date.today().weekday() == 3:
        # check in the database for last mail sent date
        try:
            is_mail_sent = Scheduled_Auto_Mail.objects.get(
                pk=1, is_sent=1, is_active=1)
            sent_date = is_mail_sent.mail_sent_date
        except Scheduled_Auto_Mail.DoesNotExist:
            sent_date = None
        now = datetime.now()
        date_string = now.strftime("%Y-%m-%d")
        if sent_date == date_string:
            print("***** Mail already sent on ", sent_date, " *****")
            pass
        else:
            a = Cron()
            a.unanswered_notification()
            Scheduled_Auto_Mail.objects.get_or_create(id=1, defaults=dict(
                mail_sent_date=date_string, is_sent=1, is_active=1))
            Scheduled_Auto_Mail.objects.filter(
                is_sent=1).update(
                mail_sent_date=date_string)
            print("***** New Notification Mail sent *****")
            a.train_spam_filter()
    else:
        print("***** Mail not sent *****")


# VIEWS FUNCTIONS

@login_required
@user_passes_test(is_moderator)
def moderator_activate(request):
    """Activate the Moderator Panel."""
    request.session['MODERATOR_ACTIVATED'] = True

    next = request.GET.get('next', '')
    try:
        # Check if there is a view corresponding to this URL,
        # throw a Resolver404 Exception otherwise
        resolve(next)
        return HttpResponseRedirect(next)
    except Resolver404:
        return HttpResponseRedirect('/moderator/')


@login_required
@user_passes_test(is_moderator)
def moderator_deactivate(request):
    """Deactivate the Moderator Panel."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        request.session['MODERATOR_ACTIVATED'] = False

    next = request.GET.get('next', '')
    try:
        resolve(next)
        return HttpResponseRedirect(next)
    except Resolver404:
        return HttpResponseRedirect('/')


def home(request):
    """Render the index Page of the Website."""
    send_remider_mail()

    if request.session.get('MODERATOR_ACTIVATED', False):
        return HttpResponseRedirect('/moderator/')

    categories = FossCategory.objects.filter(hidden=False).order_by('name')
    questions = Question.objects.filter(
        is_spam=False, is_active=True, category__hidden=False,
    ).order_by('-date_created')
    context = {
        'categories': categories,
        'questions': questions,
    }
    return render(request, "website/templates/index.html", context)


def questions(request):
    """Show all the Questions posted till now with Pagination."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        return HttpResponseRedirect('/moderator/questions/')

    categories = FossCategory.objects.filter(hidden=False).order_by('name')
    questions = Question.objects.filter(
        is_spam=False, is_active=True, category__hidden=False,
    ).order_by('-date_created')
    context = {
        'categories': categories,
        'questions': questions,
    }
    return render(request, 'website/templates/questions.html', context)


def get_question(request, question_id=None, pretty_url=None):
    """Show the details of the Question, its Answers and Comments under it."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        if is_moderator(
            request.user,
            get_object_or_404(
                Question,
                id=question_id)):
            question = get_object_or_404(
                Question, id=question_id, category__hidden=False)
            answers = question.answer_set.all()
        else:
            return render(request, 'website/templates/not-authorized.html')
    else:
        # Spam Questions should be accessible to its Author only.
        question = get_object_or_404(
            Question,
            id=question_id,
            is_active=True,
            category__hidden=False)
        if question.user != request.user and question.is_spam:
            raise Http404
        answers = question.answer_set.filter(is_active=True).all()

    sub_category = True

    if question.sub_category == "" or str(question.sub_category) == 'None':
        sub_category = False

    ans_count = len(answers)  # Includes Spam Answers
    form = AnswerQuestionForm()
    thisuserupvote = question.userUpVotes.filter(id=request.user.id).count()
    thisuserdownvote = question.userDownVotes.filter(
        id=request.user.id).count()

    ans_votes = []
    for vote in answers:
        net_ans_count = vote.num_votes
        ans_votes.append([vote.userUpVotes.filter(id=request.user.id).count(
        ), vote.userDownVotes.filter(id=request.user.id).count(),
            net_ans_count])

    main_list = list(zip(answers, ans_votes))
    context = {
        'ans_count': ans_count,
        'question': question,
        'sub_category': sub_category,
        'main_list': main_list,
        'form': form,
        'thisUserUpvote': thisuserupvote,
        'thisUserDownvote': thisuserdownvote,
        'net_count': question.num_votes,
    }
    context.update(csrf(request))

    # updating views count
    if (request.user.is_anonymous):  # if no one logged in
        question.views += 1
    elif (question.userViews.filter(id=request.user.id).count() == 0):
        question.views += 1
        question.userViews.add(request.user)

    question.save()

    context['SITE_KEY'] = settings.GOOGLE_RECAPTCHA_SITE_KEY
    return render(request, 'website/templates/get-question.html', context)


# post a new question on to forums, notification is sent to mailing list
# team@fossee.in
@login_required
@user_passes_test(account_credentials_defined, login_url='/accounts/profile/')
def new_question(request):
    """Render the page to post a new question onto the forum."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        messages.error(request, "Moderators cannot Post a New Question!")
        return HttpResponseRedirect('/moderator/')

    context = {}
    context['SITE_KEY'] = settings.GOOGLE_RECAPTCHA_SITE_KEY
    # all_category = FossCategory.objects.all()

    if (request.method == 'POST'):

        form = NewQuestionForm(request.POST, request.FILES)

        if form.is_valid():

            cleaned_data = form.cleaned_data
            question = Question()
            question.user = request.user
            question.category = cleaned_data['category']
            question.sub_category = cleaned_data['tutorial']

            if ('image' in request.FILES):
                question.image = request.FILES['image']

            if (question.sub_category == "Select a Sub Category"):

                if (str(question.category) == "Scilab Toolbox"):
                    context.update(csrf(request))
                    category = request.POST.get('category', None)
                    context['category'] = category
                    context['form'] = NewQuestionForm(category=category)
                    return render(
                        request, 'website/templates/new-question.html', context)

                question.sub_category = ""

            question.title = cleaned_data['title']
            question.body = cleaned_data['body']
            question.views = 1
            question.save()

            question.userViews.add(request.user)
            if (str(question.sub_category) == 'None'):
                question.sub_category = ""
            if (predict(question.body) == "Spam"):
                question.is_spam = True
            question.notif_flag = 1
            question.save()

            # Sending email when a new question is asked
            if question.is_spam:
                send_spam_question_notification(request.user, question)
            else:
                send_question_notification(request.user, question)

            return HttpResponseRedirect('/question/{0}/'.format(question.id))

        else:
            context.update(csrf(request))
            category = request.POST.get('category', None)
            tutorial = request.POST.get('tutorial', None)
            context['category'] = category
            context['tutorial'] = tutorial
            context['form'] = form
            return render(
                request,
                'website/templates/new-question.html',
                context)

    else:
        category = request.GET.get('category')
        form = NewQuestionForm(category=category)
        context['category'] = category

    context['form'] = form
    context.update(csrf(request))
    return render(request, 'website/templates/new-question.html', context)


@login_required
@check_recaptcha
@user_passes_test(account_credentials_defined, login_url='/accounts/profile/')
def question_answer(request, question_id):
    """Post an answer to a question asked om the forum."""
    question = get_object_or_404(
        Question,
        id=question_id,
        is_active=True,
        is_spam=False,
        category__hidden=False,
        category__disabled=False)

    if request.method == 'POST':
        form = AnswerQuestionForm(request.POST, request.FILES)
        answer = Answer()
        answer.uid = request.user.id

        if form.is_valid() and request.recaptcha_is_valid:
            cleaned_data = form.cleaned_data
            body = cleaned_data['body']
            answer.body = body.splitlines()
            answer.question = question
            answer.body = body
            if ('image' in request.FILES):
                answer.image = request.FILES['image']
            if (predict(answer.body) == "Spam"):
                answer.is_spam = True
            answer.notif_flag = 1
            answer.save()

            # SENDING EMAILS AND NOTIFICATIONS ABOUT NEW ANSWER
            if answer.is_spam:
                send_spam_answer_notification(request.user, answer)
            else:
                send_answer_notification(request.user, answer)

            return HttpResponseRedirect('/question/{0}/'.format(question_id))

        else:
            if 'body' in form.errors.as_data():
                messages.error(
                    request, form.errors['body'].as_data()[0].message)
            if 'image' in form.errors.as_data():
                messages.error(
                    request, form.errors['image'].as_data()[0].message)

    return HttpResponseRedirect('/question/{0}/'.format(question_id))


@login_required
@user_passes_test(account_credentials_defined, login_url='/accounts/profile/')
def answer_comment(request, answer_id):
    """Post a comment on an answer to a question asked on the forum."""
    answer = get_object_or_404(
        Answer,
        id=answer_id,
        is_active=True,
        is_spam=False)
    question = answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if (request.method == 'POST'):
        form = AnswerCommentForm(request.POST)

        if form.is_valid():
            body = request.POST['body']
            comment = AnswerComment(
                uid=request.user.id, answer=answer, body=body)
            if (predict(comment.body) == "Spam"):
                comment.is_spam = True
            comment.notif_flag = 1
            comment.save()

            # SENDING EMAILS AND NOTIFICATIONS ABOUT NEW COMMENT
            if comment.is_spam:
                send_spam_comment_notification(request.user, comment)
            else:
                send_comment_notification(request.user, comment)

            return HttpResponseRedirect('/question/{0}/'.format(question.id))

        else:
            if 'body' in form.errors.as_data():
                messages.error(
                    request, form.errors['body'].as_data()[0].message)
            return HttpResponseRedirect('/question/{0}/'.format(question.id))

    return HttpResponseRedirect(
        '/question/{0}/#answer{1}'.format(question.id, answer_id))


# Edit a question on forums, notification is sent to mailing list
# team@fossee.in
@login_required
@user_passes_test(account_credentials_defined, login_url='/accounts/profile/')
def edit_question(request, question_id):
    """Edit question asked on the forum."""
    context = {}
    user = request.user
    context['SITE_KEY'] = settings.GOOGLE_RECAPTCHA_SITE_KEY
    # all_category = FossCategory.objects.all()
    question = get_object_or_404(
        Question,
        id=question_id,
        is_active=True,
        category__hidden=False,
        category__disabled=False)

    # To prevent random user from manually entering the link and editing
    if ((request.user.id != question.user.id or question.answer_set
         .filter(is_active=True).count() > 0) and (
            not is_moderator(request.user, question) or not
            request.session.get('MODERATOR_ACTIVATED', False))):
        return render(request, 'website/templates/not-authorized.html')

    if (request.method == 'POST'):

        previous_title = question.title
        form = NewQuestionForm(request.POST, request.FILES)
        question.title = ''  # To prevent same title error in form
        question.save()

        if form.is_valid():

            cleaned_data = form.cleaned_data
            question.category = cleaned_data['category']
            question.sub_category = cleaned_data['tutorial']

            if ('image' in request.FILES):
                question.image = request.FILES['image']

            if (question.sub_category == "Select a Sub Category"):
                if (str(question.category) == "Scilab Toolbox"):
                    context.update(csrf(request))
                    category = request.POST.get('category', None)
                    tutorial = request.POST.get('tutorial', None)
                    context['category'] = category
                    context['tutorial'] = tutorial
                    context['form'] = form
                    return render(
                        request, 'website/templates/edit-question.html',
                        context)

                question.sub_category = ""

            change_spam = question.is_spam

            question.title = cleaned_data['title']
            question.body = cleaned_data['body']
            question.is_spam = cleaned_data['is_spam']

            if question.is_spam != change_spam:
                process_Spam(question.body, question.is_spam)

            question.views = 1
            question.save()
            question.userViews.add(request.user)
            if str(question.sub_category) == 'None':
                question.sub_category = ""
            if (not request.session.get('MODERATOR_ACTIVATED', False)):
                if (predict(question.body) == "Spam"):
                    question.is_spam = True

            question.save()

            # Sending Notifications
            if question.notif_flag == 0:
                # A Non-Spam Question is Edited
                if not (
                    request.session.get(
                        'MODERATOR_ACTIVATED',
                        False) and question.is_spam):
                    # Question not marked as spam by a Moderator during edit
                    question.notif_flag = 2
                    question.save()

                if question.is_spam:
                    send_spam_question_notification(request.user, question)
                else:
                    send_question_notification(request.user, question,
                                               previous_title=previous_title)

            else:
                # A new or recently edited Question marked as Spam is Edited
                # (notifications for it as a new/edited question were not sent)
                # If question is still spam, do nothing.
                if not question.is_spam:
                    if question.user != request.user:
                        # Send Approval Notification to Author
                        send_question_approve_notification(question)
                    # Send pending Notifications (by the name of author)
                    send_question_notification(question.user, question,
                                               previous_title=previous_title)

            return HttpResponseRedirect('/question/{0}/'.format(question.id))

        else:
            context.update(csrf(request))
            category = request.POST.get('category', None)
            tutorial = request.POST.get('tutorial', None)
            context['category'] = category
            context['tutorial'] = tutorial
            context['form'] = form
            return render(
                request,
                'website/templates/edit-question.html',
                context)

    else:
        form = NewQuestionForm(instance=question)

    context['form'] = form
    context.update(csrf(request))
    return render(request, 'website/templates/edit-question.html', context)


@login_required
def answer_update(request):
    """Update the Answer and send emails to the concerned."""
    if request.method == 'POST':
        aid = request.POST['answer_id']
        body = request.POST['answer_body']
        answer = get_object_or_404(Answer, pk=aid, is_active=True)
        question = answer.question

        if question.category.hidden or question.category.disabled:
            raise Http404

        # Answer comments excluding author's
        comments = AnswerComment.objects.filter(
            answer=answer, is_active=True).exclude(uid=answer.uid)

        if ((is_moderator(request.user,
                          question) and request.session.get(
            'MODERATOR_ACTIVATED',
            False)) or
                (request.user.id == answer.uid and not comments.exists())):
            answer.body = str(body)
            answer.is_spam = False
            if (not request.session.get('MODERATOR_ACTIVATED', False)
                    and predict(answer.body) == "Spam"):
                answer.is_spam = True
                process_Spam(answer.body,answer.is_spam)
            answer.save()

            # SENDING NOTIFICATIONS
            if answer.notif_flag == 0:
                # A Non-Spam Answer or an Answer marked as Spam by a Moderator
                # is Edited (no Notifications pending)
                answer.notif_flag = 2
                answer.save()

                if answer.is_spam:
                    send_spam_answer_notification(request.user, answer)
                else:
                    send_answer_notification(request.user, answer)
            else:
                # A new or recently edited answer marked as Spam is Edited
                # (notifications for it as a new/edited answer were not sent)
                # If answer is still spam, do nothing.
                if not answer.is_spam:
                    if answer.uid != request.user.id:
                        # Send Approval Notification to Author
                        send_answer_approve_notification(answer)
                    # Send pending Notifications (by the name of author)
                    send_answer_notification(answer.user(), answer)

            messages.success(request, "Answer is Successfully Saved!")
            return HttpResponseRedirect('/question/{0}/'.format(question.id))
        else:
            messages.error(request, "Failed to Update Answer!")
            return HttpResponseRedirect('/question/{0}/'.format(question.id))

    return render(request, 'website/templates/get-requests-not-allowed.html')


@login_required
def answer_comment_update(request):
    """Update the comment and send emails to the concerned."""
    if request.method == 'POST':
        cid = request.POST['comment_id']
        body = request.POST['comment_body']
        comment = get_object_or_404(AnswerComment, pk=cid, is_active=True)
        question = comment.answer.question

        if question.category.hidden or question.category.disabled:
            raise Http404

        if ((is_moderator(request.user, question)
             and request.session.get('MODERATOR_ACTIVATED', False)) or
                (request.user.id == comment.uid
                 and can_delete_comment(comment.answer, cid))):
            comment.body = str(body)
            comment.is_spam = False
            if (not request.session.get('MODERATOR_ACTIVATED', False) and
                    predict(comment.body) == "Spam"):
                comment.is_spam = True
                process_Spam(comment.body,comment.is_spam)
            comment.save()

            # SENDING NOTIFICATIONS
            if comment.notif_flag == 0:
                # A Non-Spam comment or an comment marked as Spam by a Moderator
                # is Edited (no Notifications pending)
                comment.notif_flag = 2
                comment.save()

                if comment.is_spam:
                    send_spam_comment_notification(request.user, comment)
                else:
                    send_comment_notification(request.user, comment)
            else:
                # A new or recently edited comment marked as Spam is Edited
                # (notifications for it as a new/edited comment were not sent)
                # If comment is still spam, do nothing.
                if not comment.is_spam:
                    if comment.uid != request.user.id:
                        # Send Approval Notification to Author
                        send_comment_approve_notification(comment)
                    # Send pending Notifications (by the name of author)
                    send_comment_notification(comment.user(), comment)

            messages.success(request, "Comment is Successfully Saved!")
            return HttpResponseRedirect('/question/{0}/'.format(question.id))
        else:
            messages.error(request, "Failed to Update Comment!")
            return HttpResponseRedirect('/question/{0}/'.format(question.id))
    return render(request, 'website/templates/get-requests-not-allowed.html')


# View for deleting question, notification is sent to mailing list
# team@fossee.in
@login_required
def question_delete(request, question_id):
    """Delete question asked on the forum."""
    question = get_object_or_404(
        Question,
        id=question_id,
        is_active=True,
        category__hidden=False,
        category__disabled=False)

    # To prevent random user from manually entering the link and deleting
    if ((request.user.id != question.user.id
         or question.answer_set.filter(is_active=True).count() > 0)
        and (
            not is_moderator(request.user, question)
            or not request.session.get('MODERATOR_ACTIVATED', False))):
        return render(request, 'website/templates/not-authorized.html')

    if (request.method == "POST"):
        question.is_active = False
        question.notif_flag = 3
        question.save()

        # SENDING NOTIFICATIONS
        delete_reason = request.POST.get('deleteQuestion', None)
        send_question_notification(request.user, question,
                                   delete_reason=delete_reason)

        return render(request, 'website/templates/question-delete.html',
                      {'title': question.title})

    # Question can only be deleted by sending POST requests
    # and not by GET requests (directly accessing the link)
    return render(request, 'website/templates/get-requests-not-allowed.html')


# View for deleting answer, notification is sent to person who posted answer
@login_required
def answer_delete(request, answer_id):
    """Delete an answer."""
    answer = get_object_or_404(Answer, id=answer_id, is_active=True)
    question = answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    # Answer comments excluding author's
    comments = AnswerComment.objects.filter(
        answer=answer, is_active=True).exclude(uid=answer.uid)

    # The second statement in if condition excludes comments made by Answer's
    # author.
    if ((request.user.id != answer.uid or comments.exists())
        and (not is_moderator(
            request.user, question) or
            not request.session.get('MODERATOR_ACTIVATED', False))):
        return render(request, 'website/templates/not-authorized.html')

    if (request.method == "POST"):
        answer.is_active = False
        answer.notif_flag = 3
        answer.save()

        # Delete all Notifications related to it
        notifications = Notification.objects.filter(aid=answer.id)
        notifications.delete()

        # Sending Emails for Answer Delete
        delete_reason = request.POST.get('deleteAnswer', None)
        send_answer_notification(request.user, answer,
                                 delete_reason=delete_reason)

        messages.success(request, "Answer Deleted Successfully!")
        return HttpResponseRedirect('/question/{0}/'.format(question.id))

    # Answer can only be deleted by sending POST requests
    # and not by GET requests (directly accessing the link)
    return render(request, 'website/templates/get-requests-not-allowed.html')


@login_required
def comment_delete(request, comment_id):
    """Delete the comment and send emails to the concerned."""
    comment = get_object_or_404(AnswerComment, pk=comment_id)
    question = comment.answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if ((request.user.id != comment.uid
         or not can_delete_comment(comment.answer, comment_id)) and (
            not is_moderator(request.user, question)
            or not request.session.get('MODERATOR_ACTIVATED', False))):
        return render(request, 'website/templates/not-authorized.html')

    if request.method == 'POST':
        comment.is_active = False
        comment.notif_flag = 3
        comment.save()

        # Delete all Notifications related to it
        notifications = Notification.objects.filter(cid=comment.id)
        notifications.delete()

        # Sending notifications
        delete_reason = request.POST.get('deleteComment', None)
        send_comment_notification(request.user, comment,
                                  delete_reason=delete_reason)

        messages.success(request, "Comment Deleted Successfully!")
        return HttpResponseRedirect('/question/{0}/'.format(question.id))

    # Comment can only be deleted by sending POST requests
    # and not by GET requests (directly accessing the link)
    return render(request, 'website/templates/get-requests-not-allowed.html')


@login_required
@user_passes_test(is_moderator)
def question_restore(request, question_id):
    """Restore a Question."""
    question = get_object_or_404(
        Question,
        id=question_id,
        is_active=False,
        category__hidden=False,
        category__disabled=False)

    if not is_moderator(
        request.user,
        question) or not request.session.get(
        'MODERATOR_ACTIVATED',
            False):
        return render(request, 'website/templates/not-authorized.html')

    question.is_active = True
    question.save()

    messages.success(request, "Question Restored Successfully!")
    return HttpResponseRedirect('/question/{0}/'.format(question_id))


@login_required
@user_passes_test(is_moderator)
def answer_restore(request, answer_id):
    """Restore an Answer."""
    answer = get_object_or_404(Answer, id=answer_id, is_active=False)
    question = answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if not is_moderator(
        request.user,
        question) or not request.session.get(
        'MODERATOR_ACTIVATED',
            False):
        return render(request, 'website/templates/not-authorized.html')

    if not question.is_active:
        messages.error(
            request,
            "Answer can only be restored when its question is not deleted.")
        return HttpResponseRedirect('/question/{0}/'.format(question.id))

    answer.is_active = True
    answer.save()

    messages.success(request, "Answer Restored Successfully!")
    return HttpResponseRedirect(
        '/question/{0}/#answer{1}'.format(question.id, answer.id))


@login_required
@user_passes_test(is_moderator)
def comment_restore(request, comment_id):
    """Restore a Comment."""
    comment = get_object_or_404(
        AnswerComment, id=comment_id, is_active=False)
    question = comment.answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if (not is_moderator(request.user, question) or
            not request.session.get('MODERATOR_ACTIVATED', False)):
        return render(request, 'website/templates/not-authorized.html')

    if not comment.answer.is_active:
        messages.error(
            request,
            "Comment can only be restored when its answer is not deleted.")
        return HttpResponseRedirect('/question/{0}/'.format(question.id))

    comment.is_active = True
    comment.save()

    messages.success(request, "Comment Restored Successfully!")
    return HttpResponseRedirect(
        '/question/{0}/#comm{1}'.format(question.id, comment.id))


# View to approve question marked as spam
@login_required
@user_passes_test(is_moderator)
def approve_spam_question(request, question_id):
    question = get_object_or_404(
        Question, id=question_id, is_active=True, is_spam=True,
        category__hidden=False, category__disabled=False)

    if (is_moderator(request.user, question) and
            request.session.get('MODERATOR_ACTIVATED', False)):
        question.is_spam = False
        question.save()
        process_Spam(question.body,question.is_spam)
        # Send Approval Notification to Author
        send_question_approve_notification(question)
        # Send Pending Notifications
        send_question_notification(question.user, question,
                                   previous_title=question.title)
        messages.success(request, "Question marked successfully as Not-Spam!")
    return HttpResponseRedirect('/question/{0}/'.format(question_id))


# View to mark answer as spam/non-spam
@login_required
@user_passes_test(is_moderator)
def mark_answer_spam(request, answer_id):
    """Mark/Unmark an Answer as a spam."""
    answer = get_object_or_404(Answer, id=answer_id, is_active=True)
    question = answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if request.method == "POST":
        choice = request.POST['selector']
        if choice == "spam":
            if not answer.is_spam:
                answer.is_spam = True
                answer.save()
                process_Spam(answer.body, answer.is_spam)
                # Send Spam Classification Notification to Author
                send_spam_answer_notification(request.user, answer)
                messages.success(
                    request, "Answer marked successfully as Spam!")
        else:
            if answer.is_spam:
                answer.is_spam = False
                answer.save()
                process_Spam(answer.body, answer.is_spam)
                # Send Approval Notification to Author
                send_answer_approve_notification(answer)
                # Send Pending Notifications (by the name of author)
                send_answer_notification(answer.user(), answer)
                messages.success(
                    request, "Answer marked successfully as Not-Spam!")
    return HttpResponseRedirect(
        '/question/{0}/#answer{1}'.format(question.id, answer.id))


# View to mark comment as spam/non-spam
@login_required
@user_passes_test(is_moderator)
def mark_comment_spam(request, comment_id):
    """Mark/Unmark an Comment as a spam."""
    comment = get_object_or_404(AnswerComment, id=comment_id, is_active=True)
    question = comment.answer.question

    if question.category.hidden or question.category.disabled:
        raise Http404

    if request.method == "POST":
        choice = request.POST['choice']
        if choice == "spam":
            comment.is_spam = True
            comment.save()
            process_Spam(comment.body, comment.is_spam)
            # Send Spam Classification Notification to Author
            send_spam_comment_notification(request.user, comment)
            messages.success(request, "Comment marked successfully as Spam!")
        else:
            comment.is_spam = False
            comment.save()
            process_Spam(comment.body, comment.is_spam)
            # Send Approval Notification to Author
            send_comment_approve_notification(comment)
            # Send Pending Notifications (by the name of author)
            send_comment_notification(comment.user(), comment)
            messages.success(
                request, "Comment marked successfully as Not-Spam!")
    return HttpResponseRedirect(
        '/question/{0}/#comm{1}'.format(question.id, comment.id))


def search(request):
    """Render 'Search Questions by Category' Page."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        messages.error(request, "Moderators cannot access the Search Page!")
        return HttpResponseRedirect('/moderator/')

    categories = FossCategory.objects.filter(hidden=False).order_by('name')
    context = {
        'categories': categories,
    }
    return render(request, 'website/templates/search.html', context)


def filter(request, category=None, tutorial=None):
    """Filter Questions based on the category and
       tutorial (sub-category) provided as arguments."""
    if category and tutorial:
        questions = Question.objects.filter(
            category__name=category,
            sub_category=tutorial,
            category__hidden=False).order_by('-date_created')
    elif tutorial is None:
        questions = Question.objects.filter(
            category__name=category,
            category__hidden=False).order_by('-date_created')

    if (not request.session.get('MODERATOR_ACTIVATED', False)):
        questions = questions.filter(is_spam=False, is_active=True)

    context = {
        'questions': questions,
        'category': category,
        'tutorial': tutorial,
    }

    return render(request, 'website/templates/filter.html', context)


@login_required
@user_passes_test(account_credentials_defined, login_url='/accounts/profile/')
def user_notifications(request, user_id):
    """Display all the Notifications recieved by the user."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        messages.error(request, "Moderators cannot access the Notifications!")
        return HttpResponseRedirect('/moderator/')

    if (user_id == request.user.id):
        try:
            notifications = Notification.objects.filter(
                uid=user_id).order_by('-date_created')
            context = {
                'notifications': notifications,
            }
            return render(
                request,
                'website/templates/notifications.html',
                context)
        except BaseException:
            return HttpResponseRedirect(
                "/user/{0}/notifications/".format(request.user.id))

    else:
        return render(request, 'website/templates/not-authorized.html')


@login_required
def clear_notifications(request):
    """Delete all the Notifications recieved by the user."""
    if request.session.get('MODERATOR_ACTIVATED', False):
        messages.error(request, "Moderators cannot clear the Notifications!")
        return HttpResponseRedirect('/moderator/')

    Notification.objects.filter(uid=request.user.id).delete()
    return HttpResponseRedirect(
        "/user/{0}/notifications/".format(request.user.id))


# MODERATOR SECTION

@login_required
@user_passes_test(is_moderator)
def moderator_home(request):
    """Render Moderator Panel Home Page."""
    if not request.session.get('MODERATOR_ACTIVATED', False):
        return HttpResponseRedirect('/')

    # If user is a super moderator
    if (request.user.groups.filter(name="forum_moderator").exists()):
        questions = Question.objects.filter(
            category__hidden=False).order_by('-date_created')
        categories = FossCategory.objects.filter(hidden=False).order_by('name')

    else:
        # Finding the moderator's categories and Getting the questions related
        # to moderator's categories
        categories = []
        questions = []
        for group in request.user.groups.all():
            category = ModeratorGroup.objects.get(group=group).category
            # Don't include hidden categories
            if category.hidden:
                continue
            categories.append(category)
            questions.extend(
                Question.objects.filter(
                    category__name=category.name).order_by('-date_created'))
        questions.sort(
            key=lambda question: question.date_created,
            reverse=True)
    context = {
        'questions': questions,
        'categories': categories,
    }

    return render(request, 'website/templates/moderator/index.html', context)


@login_required
@user_passes_test(is_moderator)
def moderator_questions(request):
    """Display all the questions belonging to the Moderator's Categories."""
    if not request.session.get('MODERATOR_ACTIVATED', False):
        return HttpResponseRedirect('/questions/')

    # If user is a super moderator
    if (request.user.groups.filter(name="forum_moderator").exists()):
        categories = FossCategory.objects.filter(hidden=False).order_by('name')
        questions = Question.objects.filter(
            category__hidden=False).order_by('-date_created')
        if ('spam' in request.GET):
            questions = questions.filter(is_spam=True)
        elif ('non-spam' in request.GET):
            questions = questions.filter(is_spam=False)

    else:
        # Finding the moderator's category questions
        questions = []
        categories = []
        for group in request.user.groups.all():
            category = ModeratorGroup.objects.get(group=group).category
            # Don't include hidden categories
            if category.hidden:
                continue
            categories.append(category)
            questions_to_add = Question.objects.filter(
                category__name=category.name).order_by('-date_created')
            if ('spam' in request.GET):
                questions_to_add = questions_to_add.filter(is_spam=True)
            elif ('non-spam' in request.GET):
                questions_to_add = questions_to_add.filter(is_spam=False)
            questions.extend(questions_to_add)
        questions.sort(
            key=lambda question: question.date_created,
            reverse=True,
        )
    context = {
        'categories': categories,
        'questions': questions,
    }
    return render(
        request,
        'website/templates/moderator/questions.html',
        context)


@login_required
@user_passes_test(is_moderator)
def moderator_unanswered(request):
    """
    Display all the Unanswered Questions belonging to the Moderator's
    Categories.
    """
    if not request.session.get('MODERATOR_ACTIVATED', False):
        return HttpResponseRedirect('/')

    # If user is a super moderator
    if (request.user.groups.filter(name="forum_moderator").exists()):
        categories = FossCategory.objects.filter(hidden=False).order_by('name')
        questions = Question.objects.filter(
            is_active=True, category__hidden=False).order_by('-date_created')

    else:
        # Finding the moderator's category questions
        questions = []
        categories = []
        for group in request.user.groups.all():
            category = ModeratorGroup.objects.get(group=group).category
            # Don't include hidden categories
            if category.hidden:
                continue
            categories.append(category)
            questions.extend(
                Question.objects.filter(
                    category__name=category.name,
                    is_active=True).order_by('-date_created'))
        questions.sort(
            key=lambda question: question.date_created,
            reverse=True,
        )
    context = {
        'categories': categories,
        'questions': questions,
    }
    return render(
        request,
        'website/templates/moderator/unanswered.html',
        context)


@login_required
@user_passes_test(is_moderator)
def train_spam_filter(request):
    """Re-train the Spam Filter."""
    next = request.GET.get('next', '')
    train()
    try:
        resolve(next)
        return HttpResponseRedirect(next)
    except Resolver404:
        return HttpResponseRedirect('/moderator/')


# AJAX SECTION

@csrf_exempt
def ajax_tutorials(request):
    """Don't know the use :P."""
    if request.method == 'POST':
        category = request.POST.get('category')
        tutorials = SubFossCategory.objects.filter(parent_id=category)
        if tutorials.exists():
            context = {
                'tutorials': tutorials,
            }
            return render(
                request,
                'website/templates/ajax-tutorials.html',
                context)
        else:
            return HttpResponse('No sub-category in category.')
    else:
        return render(
            request,
            'website/templates/get-requests-not-allowed.html')


@login_required
@csrf_exempt
def ajax_notification_remove(request):
    """Clear (Delete) the Notification."""
    if request.method == "POST":

        nid = request.POST["notification_id"]

        try:
            notification = get_object_or_404(Notification, pk=nid)
            if (notification.uid == request.user.id):
                notification.delete()
                return HttpResponse('removed')
            else:
                return HttpResponse('Unauthorized user.')
        except BaseException:
            return HttpResponse('Notification not found.')

    else:
        return render(
            request,
            'website/templates/get-requests-not-allowed.html')


@csrf_exempt
def ajax_keyword_search(request):
    """Display the Questions based on the entered keyword."""
    if request.method == "POST":
        key = request.POST['key']
        questions = (
            Question.objects.filter(title__icontains=key).filter(
                is_spam=False,
                is_active=True,
                category__hidden=False,
            ) | Question.objects.filter(category__name=key).filter(
                is_spam=False,
                is_active=True,
                category__hidden=False,
            )
        ).distinct().order_by('-date_created')
        context = {
            'questions': questions
        }
        return render(
            request,
            'website/templates/ajax-keyword-search.html',
            context)
    else:
        return render(
            request,
            'website/templates/get-requests-not-allowed.html')


# return number of votes and initial votes
# user who asked the question,cannot vote his/or anwser,
# other users can post votes
@login_required
def ajax_vote_post(request):

    post_id = int(request.POST.get('id'))
    vote_type = request.POST.get('type')
    vote_action = request.POST.get('action')
    cur_post = get_object_or_404(Question, id=post_id, is_active=True)
    thisuserupvote = cur_post.userUpVotes.filter(
        id=request.user.id, is_active=True).count()
    thisuserdownvote = cur_post.userDownVotes.filter(
        id=request.user.id, is_active=True).count()
    initial_votes = cur_post.num_votes

    if (request.user.id != cur_post.user.id):

        # This condition is for adding vote
        if vote_action == 'vote':
            if (thisuserupvote == 0) and (thisuserdownvote == 0):
                if vote_type == 'up':
                    cur_post.userUpVotes.add(request.user)
                elif vote_type == 'down':
                    cur_post.userDownVotes.add(request.user)
                else:
                    return HttpResponse(initial_votes)
            else:
                if (thisuserupvote == 1) and (vote_type == 'down'):
                    cur_post.userUpVotes.remove(request.user)
                    cur_post.userDownVotes.add(request.user)
                elif (thisuserdownvote == 1) and (vote_type == 'up'):
                    cur_post.userDownVotes.remove(request.user)
                    cur_post.userUpVotes.add(request.user)
                else:
                    return HttpResponse(initial_votes)

        # This condition is for canceling vote
        elif vote_action == 'recall-vote':
            if (vote_type == 'up') and (thisuserupvote == 1):
                cur_post.userUpVotes.remove(request.user)
            elif (vote_type == 'down') and (thisuserdownvote == 1):
                cur_post.userDownVotes.remove(request.user)
            else:
                return HttpResponse(initial_votes)
        else:
            return HttpResponse("Error: Bad Action.")

        num_votes = cur_post.userUpVotes.count() - cur_post\
            .userDownVotes.count()
        cur_post.num_votes = num_votes
        cur_post.save()
        return HttpResponse(num_votes)

    else:
        return HttpResponse(initial_votes)


# return number of votes and initial votes
# user who posted the answer, cannot vote his/or anwser,
# other users can post votes
@login_required
def ajax_ans_vote_post(request):

    post_id = int(request.POST.get('id'))
    vote_type = request.POST.get('type')
    vote_action = request.POST.get('action')
    cur_post = get_object_or_404(Answer, id=post_id, is_active=True)
    thisuserupvote = cur_post.userUpVotes.filter(
        id=request.user.id, is_active=True).count()
    thisuserdownvote = cur_post.userDownVotes.filter(
        id=request.user.id, is_active=True).count()
    initial_votes = cur_post.num_votes

    if (request.user.id != cur_post.uid):

        # This condition is for voting
        if (vote_action == 'vote'):
            if (thisuserupvote == 0) and (thisuserdownvote == 0):
                if vote_type == 'up':
                    cur_post.userUpVotes.add(request.user)
                elif vote_type == 'down':
                    cur_post.userDownVotes.add(request.user)
                else:
                    return HttpResponse(initial_votes)
            else:
                if (thisuserupvote == 1) and (vote_type == 'down'):
                    cur_post.userUpVotes.remove(request.user)
                    cur_post.userDownVotes.add(request.user)
                elif (thisuserdownvote == 1) and (vote_type == 'up'):
                    cur_post.userDownVotes.remove(request.user)
                    cur_post.userUpVotes.add(request.user)
                else:
                    return HttpResponse(initial_votes)

        # This condition is for canceling vote
        elif (vote_action == 'recall-vote'):
            if (vote_type == 'up') and (thisuserupvote == 1):
                cur_post.userUpVotes.remove(request.user)
            elif (vote_type == 'down') and (thisuserdownvote == 1):
                cur_post.userDownVotes.remove(request.user)
            else:
                return HttpResponse(initial_votes)
        else:
            return HttpResponse(initial_votes)

        num_votes = cur_post.userUpVotes.count() - cur_post\
            .userDownVotes.count()

        cur_post.num_votes = num_votes
        cur_post.save()
        return HttpResponse(num_votes)

    else:
        return HttpResponse(initial_votes)


# ALL NOTIFICATIONS BELOW

def send_question_notification(
        user,
        question,
        previous_title=None,
        delete_reason=None):
    # Values of flag:
    # 0 --> No Pending Notifications, 1 --> Recently Added,
    # 2 --> Recently Updated,  3 --> Recently Deleted.
    flag = question.notif_flag
    # Question Added Recently
    if flag == 1:
        subject = "FOSSEE Forums - {0} - New Question".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = [question.category.email]
        html_message = render_to_string(
            'website/templates/emails/new_question_email.html', {
                'title': question.title,
                'category': question.category,
                'body': question.body,
                'link': settings.DOMAIN_NAME + '/question/' + str(question.id),
            })
        plain_message = strip_tags(html_message)
        send_email(subject, plain_message, html_message, from_email, to)
    # Question Edited Recently
    elif flag == 2:
        subject = "FOSSEE Forums - {0} - Question Edited".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        html_message = render_to_string(
            'website/templates/emails/edited_question_email.html', {
                'title': question.title,
                'previous_title': previous_title,
                'category': question.category,
                'body': question.body,
                'link': settings.DOMAIN_NAME + '/question/' + str(question.id),
            })
        plain_message = strip_tags(html_message)
        to = [question.category.email, settings.FORUM_NOTIFICATION]

        # Getting emails of everyone in Question Thread and appending in 'to'
        if question.user != user:
            mail_uids = to_uids(question)
            for uid in mail_uids:
                to.append(get_user_email(uid))

        send_email_as_to(subject, plain_message, html_message, from_email, to)
    # Question Deleted Recently
    # No need to send any notif if deleted by author
    elif flag == 3 and question.user != user:
        subject = "FOSSEE Forums - {0} - Question Deleted".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = []
        html_message = render_to_string(
            'website/templates/emails/deleted_question_email.html',
            {
                'title': question.title,
                'category': question.category,
                'body': question.body,
                'reason': delete_reason,
            })
        plain_message = strip_tags(html_message)
        # If Spam Question is deleted, send mail to question author only
        if question.is_spam:
            to = [question.user.email]
        else:
            mail_uids = to_uids(question)
            for uid in mail_uids:
                to.append(get_user_email(uid))

        send_email_as_to(subject, plain_message, html_message, from_email, to)

    question.notif_flag = 0
    question.save()


def send_answer_notification(user, answer, delete_reason=None):
    flag = answer.notif_flag
    question = answer.question
    # Answer Added Recently
    if flag == 1:
        from_email = settings.SENDER_EMAIL
        html_message = render_to_string(
            'website/templates/emails/new_answer_email.html', {
                'title': question.title,
                'category': question.category,
                'link': settings.DOMAIN_NAME + '/question/' +
                str(question.id) + "#answer" + str(answer.id),
            })
        plain_message = strip_tags(html_message)

        not_to_notify = [user.id]

        # Notifying the Question Author
        # if question.user.id not in not_to_notify and answer.is_spam == False:
        if question.user.id not in not_to_notify:
            notification = Notification(
                uid=question.user.id, qid=question.id, aid=answer.id)
            notification.save()

            subject = "FOSSEE Forums - {0} - Your question has been Answered"\
                .format(
                    question.category)
            to = [question.user.email]
            send_email(subject, plain_message, html_message, from_email, to)

            not_to_notify.append(question.user.id)

        # Email and Notification for all user in this thread
        mail_uids = to_uids(question)
        mail_uids.difference_update(set(not_to_notify))

        subject = "FOSSEE Forums - {0} - Question has been Answered".format(
            question.category)
        to = []
        for mail_uid in mail_uids:
            notification = Notification(
                uid=mail_uid, qid=question.id, aid=answer.id)
            notification.save()
            # Appending user email in 'to' list
            to.append(get_user_email(mail_uid))

        # Sending Email to everyone in 'to' list individually
        send_email_as_to(subject, plain_message, html_message, from_email, to)
    # Answer Edited Recently
    elif flag == 2:
        subject = "FOSSEE Forums - {0} - Answer Edited".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = []

        if answer.uid == user.id:
            # Answer edited by Answer Author
            html_message = render_to_string(
                'website/templates/emails/edited_answer_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#answer" + str(answer.id),
                })
            plain_message = strip_tags(html_message)

            to.append(question.user.email)
            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)
        else:
            # Answer edited by a Moderator
            html_message = render_to_string(
                'website/templates/emails/edited_answer_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'link': settings.DOMAIN_NAME + '/question/' + str(question.id)
                    + "#answer" + str(answer.id),
                    'by_moderator': True,
                })
            plain_message = strip_tags(html_message)

            mail_uids = to_uids(question)
            for uid in mail_uids:
                to.append(get_user_email(uid))

            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)
    # Answer deleted recently
    elif flag == 3:
        subject = "FOSSEE Forums - {0} - Answer Deleted".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = [question.user.email]

        if answer.uid == user.id:
            # Answer deleted by Answer Author
            # Don't send mail if spam answer is deleted by author
            if not answer.is_spam:
                html_message = render_to_string(
                    'website/templates/emails/deleted_answer_email.html',
                    {
                        'title': question.title,
                        'category': question.category,
                        'body': question.body,
                        'answer': answer.body,
                    })
                plain_message = strip_tags(html_message)

                to.append(question.category.email)
                send_email_as_to(
                    subject,
                    plain_message,
                    html_message,
                    from_email,
                    to)
        else:
            # Answer deleted by a Moderator
            html_message = render_to_string(
                'website/templates/emails/deleted_answer_email.html',
                {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'answer': answer.body,
                    'reason': delete_reason,
                    'by_moderator': True,
                })
            plain_message = strip_tags(html_message)

            if answer.is_spam:
                # Send mail to Answer Author only if spam answer deleted by
                # moderator
                to = [get_user_email(answer.uid)]
            else:
                to.append(get_user_email(answer.uid))
                for comment in AnswerComment.objects.filter(
                        answer=answer, is_active=True):
                    to.append(get_user_email(comment.uid))
                to = list(set(to))   # Removing Duplicates

            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)

    # Updating notif flag after sending mails according to the previous flag.
    answer.notif_flag = 0
    answer.save()


def send_comment_notification(user, comment, delete_reason=None):
    flag = comment.notif_flag
    answer = comment.answer
    question = comment.answer.question
    # Comment Added Recently
    if flag == 1:
        from_email = settings.SENDER_EMAIL
        not_to_notify = [user.id]

        # Notifying the Question Author
        if question.user.id not in not_to_notify:
            notification = Notification(
                uid=question.user.id,
                qid=question.id,
                aid=answer.id,
                cid=comment.id)
            notification.save()

            subject = "FOSSEE Forums - {0} - New Comment under your Question"\
                .format(
                    question.category)
            to = [question.user.email]
            html_message = render_to_string(
                'website/templates/emails/new_comment_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#comm" + str(comment.id),
                })
            plain_message = strip_tags(html_message)
            send_email(subject, plain_message, html_message, from_email, to)

            not_to_notify.append(question.user.id)

        # Notifying the Answer Author
        if answer.uid not in not_to_notify:
            notification = Notification(
                uid=answer.uid,
                qid=question.id,
                aid=answer.id,
                cid=comment.id)
            notification.save()

            subject = "FOSSEE Forums - {0} - New Comment on your answer".format(
                question.category)
            to = [get_user_email(answer.uid)]
            html_message = render_to_string(
                'website/templates/emails/new_comment_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'answer': answer.body,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#comm" + str(comment.id),
                })
            plain_message = strip_tags(html_message)
            send_email(subject, plain_message, html_message, from_email, to)

            not_to_notify.append(answer.uid)

        # Notifying the Last Comment Author
        answer_comments = AnswerComment.objects.filter(
            answer=answer, is_active=True).exclude(
            uid=user.id).order_by('-date_created')
        if answer_comments.exists(
        ) and answer_comments[0].uid not in not_to_notify:
            last_comment = answer_comments[0]

            notification = Notification(
                uid=last_comment.uid,
                qid=question.id,
                aid=answer.id,
                cid=comment.id)
            notification.save()

            subject = "FOSSEE Forums - {0} - Your Comment has a Reply".format(
                question.category)
            to = [get_user_email(last_comment.uid)]
            html_message = render_to_string(
                'website/templates/emails/new_comment_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'last_comment': last_comment.body,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#comm" + str(comment.id),
                })
            plain_message = strip_tags(html_message)
            send_email(subject, plain_message, html_message, from_email, to)

            not_to_notify.append(last_comment.uid)

        # Notifying all other users in the thread
        mail_uids = to_uids(question)
        mail_uids.difference_update(set(not_to_notify))

        subject = "FOSSEE Forums - {0} - New Comment under the Question".format(
            question.category)
        html_message = render_to_string(
            'website/templates/emails/new_comment_email.html', {
                'title': question.title,
                'category': question.category,
                'body': question.body,
                'link': settings.DOMAIN_NAME + '/question/' +
                str(question.id) + "#comm" + str(comment.id),
            })
        plain_message = strip_tags(html_message)

        to = []
        for mail_uid in mail_uids:
            notification = Notification(
                uid=mail_uid,
                qid=question.id,
                aid=answer.id,
                cid=comment.id)
            notification.save()
            # Appending user email in 'to' list
            to.append(get_user_email(mail_uid))

        # Sending Email to everyone in 'to' list individually
        send_email_as_to(subject, plain_message, html_message, from_email, to)
    # Comment Edited Recently
    if flag == 2:
        subject = "FOSSEE Forums - {0} - Comment Edited".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = []

        if comment.uid == user.id:
            # Comment edited by Comment Author
            html_message = render_to_string(
                'website/templates/emails/edited_comment_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#comm" + str(comment.id),
                })
            plain_message = strip_tags(html_message)

            to.append(get_user_email(answer.uid))
            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)
        else:
            # Comment edited by a Moderator
            html_message = render_to_string(
                'website/templates/emails/edited_comment_email.html', {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'link': settings.DOMAIN_NAME + '/question/' +
                    str(question.id) + "#comm" + str(comment.id),
                    'by_moderator': True,
                })
            plain_message = strip_tags(html_message)

            mail_uids = to_uids(question)
            mail_uids.discard(user.id)
            for uid in mail_uids:
                to.append(get_user_email(uid))

            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)

    # Comment Deleted Recently
    elif flag == 3:
        subject = "FOSSEE Forums - {0} - Comment Deleted".format(
            question.category)
        from_email = settings.SENDER_EMAIL
        to = []

        if comment.uid == user.id:
            # Comment Deleted by Comment Author
            # Don't send mail if spam comment is deleted by author
            if not comment.is_spam:
                html_message = render_to_string(
                    'website/templates/emails/deleted_comment_email.html',
                    {
                        'title': question.title,
                        'category': question.category,
                        'body': question.body,
                        'answer': answer.body,
                        'comment': comment.body,
                    })
                plain_message = strip_tags(html_message)

                to.append(get_user_email(answer.uid))
                send_email_as_to(
                    subject,
                    plain_message,
                    html_message,
                    from_email,
                    to)

        else:
            # Comment deleted by a Moderator
            html_message = render_to_string(
                'website/templates/emails/deleted_comment_email.html',
                {
                    'title': question.title,
                    'category': question.category,
                    'body': question.body,
                    'answer': answer.body,
                    'comment': comment.body,
                    'reason': delete_reason,
                    'by_moderator': True,
                })
            plain_message = strip_tags(html_message)

            if comment.is_spam:
                # Send mail to Comment Author only if spam comment deleted by
                # moderator
                to = [get_user_email(comment.uid)]
            else:
                mail_uids = to_uids(question)
                for uid in mail_uids:
                    to.append(get_user_email(uid))

            send_email_as_to(
                subject,
                plain_message,
                html_message,
                from_email,
                to)

    # Updating notif flag after sending mails according to the previous flag.
    comment.notif_flag = 0
    comment.save()


def send_spam_question_notification(user, question):
    subject = "FOSSEE Forums - {0} - Question Classified as Spam".format(
        question.category)
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/spam_question_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'link': settings.DOMAIN_NAME + '/question/' + str(question.id),
        })
    plain_message = strip_tags(html_message)

    if question.user != user:
        # Question marked as spam by a Moderator
        to = [question.user.email]
    else:
        # Question marked as spam during interaction by Author
        # (by the spamFilter), send notification to Moderators
        to = []
        uids = mod_uids(question)
        for uid in uids:
            to.append(get_user_email(uid))

    send_email_as_to(subject, plain_message, html_message, from_email, to)


def send_spam_answer_notification(user, answer):
    question = answer.question

    subject = "FOSSEE Forums - {0} - Answer Classified as Spam".format(
        question.category)
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/spam_answer_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'answer': answer.body,
            'link': settings.DOMAIN_NAME + '/question/' +
            str(question.id) + "#answer" + str(answer.id),
        })
    plain_message = strip_tags(html_message)

    if answer.uid != user.id:
        # Answer marked as spam by a Moderator
        to = [get_user_email(answer.uid)]
    else:
        # Answer marked as spam during interaction by Author
        # (by the spamFilter), send notification to Moderators
        to = []
        uids = mod_uids(question)
        for uid in uids:
            to.append(get_user_email(uid))

    send_email_as_to(subject, plain_message, html_message, from_email, to)


def send_spam_comment_notification(user, comment):
    answer = comment.answer
    question = answer.question

    subject = "FOSSEE Forums - {0} - Comment Classified as Spam".format(
        question.category)
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/spam_comment_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'answer': answer.body,
            'comment': comment.body,
            'link': settings.DOMAIN_NAME + '/question/' +
            str(question.id) + "#comm" + str(comment.id),
        })
    plain_message = strip_tags(html_message)

    if comment.uid != user.id:
        # Comment marked as spam by a Moderator
        to = [get_user_email(comment.uid)]
    else:
        # Comment marked as spam during interaction by Author
        # (by the spamFilter), send notification to Moderators
        to = []
        uids = mod_uids(question)
        for uid in uids:
            to.append(get_user_email(uid))

    send_email_as_to(subject, plain_message, html_message, from_email, to)


def send_question_approve_notification(question):
    subject = "FOSSEE Forums - {0} - Question Approved".format(
        question.category)
    to = [question.user.email]
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/approved_question_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'link': settings.DOMAIN_NAME + '/question/' + str(question.id),
        })
    plain_message = strip_tags(html_message)
    send_email(subject, plain_message, html_message, from_email, to)


def send_answer_approve_notification(answer):
    question = answer.question

    subject = "FOSSEE Forums - {0} - Answer Approved".format(question.category)
    to = [get_user_email(answer.uid)]
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/approved_answer_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'answer': answer.body,
            'link': settings.DOMAIN_NAME + '/question/' +
            str(question.id) + "#answer" + str(answer.id),
        })
    plain_message = strip_tags(html_message)
    send_email(subject, plain_message, html_message, from_email, to)


def send_comment_approve_notification(comment):
    answer = comment.answer
    question = answer.question

    subject = "FOSSEE Forums - {0} - Comment Approved".format(
        question.category)
    to = [get_user_email(comment.uid)]
    from_email = settings.SENDER_EMAIL
    html_message = render_to_string(
        'website/templates/emails/approved_comment_email.html', {
            'title': question.title,
            'category': question.category,
            'body': question.body,
            'answer': answer.body,
            'comment': comment.body,
            'link': settings.DOMAIN_NAME + '/question/' +
            str(question.id) + "#comm" + str(comment.id),
        })
    plain_message = strip_tags(html_message)
    send_email(subject, plain_message, html_message, from_email, to)
